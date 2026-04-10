from __future__ import annotations

import argparse
import re
from collections import defaultdict
from copy import copy
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_LOW_FILE = BASE_DIR / "low.xlsx"
DEFAULT_TEMPLATE_FILE = BASE_DIR / "worklog_set.xlsx"
DEFAULT_OUTPUT_DIR = BASE_DIR / "output"

WORKLOG_SHEET_NAME = "4월"
LOW_HEADER_ROW = 5
LOW_WEEKDAY_ROW = 6
LOW_DATA_START_ROW = 7
WORKLOG_ITEM_COL = 2
WORKLOG_DAILY_COUNT_COL = 5
WORKLOG_DAILY_PEOPLE_COL = 6
WORKLOG_MONTH_COUNT_COL = 7
WORKLOG_MONTH_PEOPLE_COL = 8

SKIP_WORKLOG_ITEMS = {
    "(이용상담)문자",
    "(이용상담)일반",
    "수강신청",
    "개강 오리엔테이션",
    "작품전시회",
    "평생즐기제",
    "특강 및 외부대회 참여지원",
    "지역탐방",
    "아카데미",
    "스승의날 행사",
    "반장/강사 간담회",
    "모니터링",
    "강사평가",
    "강사 만족도조사",
    "프로그램 만족도 조사",
}

MANUAL_PROGRAM_MAP = {
    "한글교실(초급1)": ["한글교실(초급1반)"],
    "한글교실(초급2)": ["한글교실(초급2반)"],
    "한글서예1반": ["한글서예1반(화)"],
    "한글서예2반": ["한글서예2반(수)"],
    "한글서예3반": ["한글서예3반(목)"],
    "한문서예1반": ["한문서예1반(월)"],
    "한문서예2반": ["한문서예2반(수)"],
    "한문서예3반": ["한문서예3반(월)"],
}


def normalize_text(value: object) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_key(value: object) -> str:
    text = normalize_text(value)
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"\([^)]*\)$", "", text)
    text = text.replace("반", "")
    return text


def excel_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def find_date_columns(low_sheet) -> dict[date, int]:
    columns: dict[date, int] = {}
    for col in range(1, low_sheet.max_column + 1):
        day = excel_date(low_sheet.cell(LOW_HEADER_ROW, col).value)
        if day is not None:
            columns[day] = col
    return columns


def latest_available_date(date_columns: dict[date, int]) -> date:
    if not date_columns:
        raise ValueError("low.xlsx에서 날짜 컬럼을 찾지 못했습니다.")
    return max(date_columns)


def latest_populated_date(low_sheet, date_columns: dict[date, int]) -> date:
    populated_dates = []
    for day, col in date_columns.items():
        has_attendance = any(
            low_sheet.cell(row, col).value not in (None, "")
            for row in range(LOW_DATA_START_ROW, low_sheet.max_row + 1)
        )
        if has_attendance:
            populated_dates.append(day)

    if not populated_dates:
        return latest_available_date(date_columns)

    return max(populated_dates)


def build_attendance_index(low_sheet, date_columns: dict[date, int]):
    daily_counts: dict[str, dict[date, int]] = defaultdict(lambda: defaultdict(int))
    available_programs: set[str] = set()

    for row in range(LOW_DATA_START_ROW, low_sheet.max_row + 1):
        program = normalize_text(low_sheet.cell(row, 2).value)
        if not program:
            continue

        available_programs.add(program)

        for day, col in date_columns.items():
            value = low_sheet.cell(row, col).value
            if value not in (None, ""):
                daily_counts[program][day] += 1

    return daily_counts, available_programs


def match_programs(worklog_item: str, available_programs: set[str]) -> list[str]:
    if worklog_item in MANUAL_PROGRAM_MAP:
        return [name for name in MANUAL_PROGRAM_MAP[worklog_item] if name in available_programs]

    if worklog_item in available_programs:
        return [worklog_item]

    item_key = normalize_key(worklog_item)
    matches = []
    for program in available_programs:
        program_key = normalize_key(program)
        if item_key and (item_key == program_key or item_key in program_key or program_key in item_key):
            matches.append(program)

    return sorted(matches)


def month_to_date_days(date_columns: dict[date, int], target_date: date) -> list[date]:
    return sorted(
        day
        for day in date_columns
        if day.year == target_date.year and day.month == target_date.month and day <= target_date
    )


def set_output_cell(cell, value: int) -> None:
    cell.value = value if value else None


def copy_template_date_style(sheet, target_date: date) -> None:
    cell = sheet["A3"]
    cell.value = datetime(target_date.year, target_date.month, target_date.day)
    cell.number_format = 'yyyy"년" m"월" d"일" dddd'


def fill_worklog(low_file: Path, template_file: Path, output_file: Path, target_date: date | None) -> dict:
    low_wb = load_workbook(low_file, data_only=True)
    low_sheet = low_wb.active
    date_columns = find_date_columns(low_sheet)
    selected_date = target_date or latest_populated_date(low_sheet, date_columns)

    if selected_date not in date_columns:
        available = ", ".join(day.isoformat() for day in sorted(date_columns))
        raise ValueError(f"{selected_date.isoformat()} 날짜가 low.xlsx에 없습니다. 사용 가능 날짜: {available}")

    daily_counts, available_programs = build_attendance_index(low_sheet, date_columns)
    month_days = month_to_date_days(date_columns, selected_date)

    template_wb = load_workbook(template_file)
    if WORKLOG_SHEET_NAME not in template_wb.sheetnames:
        raise ValueError(f"{template_file.name}에 '{WORKLOG_SHEET_NAME}' 시트가 없습니다.")

    sheet = template_wb[WORKLOG_SHEET_NAME]
    copy_template_date_style(sheet, selected_date)

    matched_rows = []
    skipped_rows = []
    unmatched_rows = []

    for row in range(7, sheet.max_row + 1):
        item = normalize_text(sheet.cell(row, WORKLOG_ITEM_COL).value)
        if not item:
            continue

        if item in SKIP_WORKLOG_ITEMS:
            skipped_rows.append({"row": row, "item": item, "reason": "출결 원본 대상 아님"})
            continue

        programs = match_programs(item, available_programs)
        if not programs:
            unmatched_rows.append({"row": row, "item": item})
            continue

        daily_people = sum(daily_counts[program].get(selected_date, 0) for program in programs)
        daily_sessions = sum(1 for program in programs if daily_counts[program].get(selected_date, 0) > 0)
        month_people = sum(sum(daily_counts[program].get(day, 0) for day in month_days) for program in programs)
        month_sessions = sum(
            1
            for program in programs
            for day in month_days
            if daily_counts[program].get(day, 0) > 0
        )

        set_output_cell(sheet.cell(row, WORKLOG_DAILY_COUNT_COL), daily_sessions)
        set_output_cell(sheet.cell(row, WORKLOG_DAILY_PEOPLE_COL), daily_people)
        set_output_cell(sheet.cell(row, WORKLOG_MONTH_COUNT_COL), month_sessions)
        set_output_cell(sheet.cell(row, WORKLOG_MONTH_PEOPLE_COL), month_people)

        matched_rows.append(
            {
                "row": row,
                "item": item,
                "programs": programs,
                "daily_sessions": daily_sessions,
                "daily_people": daily_people,
                "month_sessions": month_sessions,
                "month_people": month_people,
            }
        )

    output_file.parent.mkdir(parents=True, exist_ok=True)
    template_wb.save(output_file)

    return {
        "target_date": selected_date.isoformat(),
        "output_file": str(output_file),
        "matched_rows": matched_rows,
        "skipped_rows": skipped_rows,
        "unmatched_rows": unmatched_rows,
    }


def write_report(summary: dict, report_file: Path) -> None:
    lines = [
        "# 업무일지 자동화 실행 리포트",
        "",
        f"- 기준일: {summary['target_date']}",
        f"- 결과 파일: {summary['output_file']}",
        f"- 자동 매칭 행: {len(summary['matched_rows'])}",
        f"- 제외 행: {len(summary['skipped_rows'])}",
        f"- 미매칭 행: {len(summary['unmatched_rows'])}",
        "",
        "## 자동 매칭",
        "",
        "| 행 | 업무일지 항목 | 원본 프로그램 | 일계 건 | 일계 명 | 월계 건 | 월계 명 |",
        "|---:|---|---|---:|---:|---:|---:|",
    ]

    for row in summary["matched_rows"]:
        programs = ", ".join(row["programs"])
        lines.append(
            "| {row} | {item} | {programs} | {daily_sessions} | {daily_people} | {month_sessions} | {month_people} |".format(
                row=row["row"],
                item=row["item"],
                programs=programs,
                daily_sessions=row["daily_sessions"],
                daily_people=row["daily_people"],
                month_sessions=row["month_sessions"],
                month_people=row["month_people"],
            )
        )

    lines.extend(["", "## 미매칭", ""])
    if summary["unmatched_rows"]:
        for row in summary["unmatched_rows"]:
            lines.append(f"- {row['row']}행: {row['item']}")
    else:
        lines.append("- 없음")

    lines.extend(["", "## 제외", ""])
    for row in summary["skipped_rows"]:
        lines.append(f"- {row['row']}행: {row['item']} ({row['reason']})")

    report_file.write_text("\n".join(lines) + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="low.xlsx 출결 데이터를 업무일지 양식에 입력합니다.")
    parser.add_argument("--date", help="기준일. 예: 2026-04-11")
    parser.add_argument("--low", default=str(DEFAULT_LOW_FILE), help="출결 원본 xlsx 경로")
    parser.add_argument("--template", default=str(DEFAULT_TEMPLATE_FILE), help="업무일지 템플릿 xlsx 경로")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="결과 저장 폴더")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    target_date = date.fromisoformat(args.date) if args.date else None
    selected_date_text = args.date or "latest"

    output_dir = Path(args.output_dir)
    output_file = output_dir / f"worklog_result_{selected_date_text}.xlsx"
    report_file = output_dir / f"worklog_report_{selected_date_text}.md"

    summary = fill_worklog(
        low_file=Path(args.low),
        template_file=Path(args.template),
        output_file=output_file,
        target_date=target_date,
    )

    if selected_date_text == "latest":
        actual_date = summary["target_date"]
        output_file = output_dir / f"worklog_result_{actual_date}.xlsx"
        report_file = output_dir / f"worklog_report_{actual_date}.md"
        Path(summary["output_file"]).replace(output_file)
        summary["output_file"] = str(output_file)

    write_report(summary, report_file)

    print(f"기준일: {summary['target_date']}")
    print(f"결과 파일: {output_file}")
    print(f"리포트 파일: {report_file}")
    print(f"자동 매칭: {len(summary['matched_rows'])}개")
    print(f"미매칭: {len(summary['unmatched_rows'])}개")


if __name__ == "__main__":
    main()
