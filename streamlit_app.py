from __future__ import annotations

import re
import tempfile
import zipfile
from datetime import date
from io import BytesIO
from pathlib import Path

import streamlit as st
from openpyxl import load_workbook

from automate_worklog import fill_worklog_dates, find_date_columns, latest_populated_date


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE_FILES = [
    ("worklog_set1", BASE_DIR / "worklog_set1.xlsx"),
    ("worklog_set2", BASE_DIR / "worklog_set2.xlsx"),
]
PHONE_PATTERN = re.compile(
    r"(?<!\d)(?:01[016789][-\s.]?\d{3,4}[-\s.]?\d{4}|0(?:2|[3-6][1-5]|70|50[2-8])[-\s.]?\d{3,4}[-\s.]?\d{4})(?!\d)"
)


st.set_page_config(
    page_title="업무일지 자동 생성",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
      html, body, [class*="css"] {
        font-family: "Inter", "Segoe UI", "Malgun Gothic", Arial, sans-serif;
      }

      .stApp {
        background: #f7f8fa;
      }

      [data-testid="stHeader"] {
        background: rgba(247, 248, 250, 0.94);
      }

      .block-container {
        max-width: 940px;
        padding-top: 42px;
        padding-bottom: 56px;
      }

      .hero {
        margin-bottom: 28px;
      }

      .hero h1 {
        margin: 0 0 12px;
        color: #0f172a;
        font-size: 32px;
        font-weight: 800;
        line-height: 1.3;
      }

      .hero p {
        margin: 0;
        max-width: 640px;
        color: #475569;
        font-size: 15px;
        line-height: 1.6;
      }

      .eyebrow {
        margin: 0 0 8px;
        color: #0d9488;
        font-size: 13px;
        font-weight: 800;
        letter-spacing: 0.5px;
      }

      .section-title {
        margin: 0 0 6px;
        color: #1e293b;
        font-size: 18px;
        font-weight: 700;
      }

      .section-note {
        margin: 0 0 12px;
        color: #64748b;
        font-size: 14px;
        line-height: 1.5;
      }

      .field-label {
        margin: 0 0 8px;
        color: #1e293b;
        font-size: 16px;
        font-weight: 700;
        line-height: 1.4;
      }

      .privacy-note {
        border: 1px solid #e2e8f0;
        border-radius: 8px;
        background: #f8fafc;
        padding: 12px 14px;
        color: #334155;
        font-size: 14px;
        line-height: 1.5;
      }


      .result-panel {
        margin-top: 18px;
        border: 1px solid #ccfbf1;
        border-radius: 8px;
        background: #f0fdfa;
        padding: 16px;
      }

      .result-panel h3 {
        margin: 0 0 6px;
        color: #115e59;
        font-size: 18px;
        font-weight: 800;
      }

      .result-panel p {
        margin: 0 0 10px;
        color: #134e4a;
        font-size: 14px;
      }

      .file-list {
        margin: 0;
        padding-left: 18px;
        color: #0f766e;
        font-size: 14px;
        line-height: 1.7;
      }

      div[data-testid="stMetric"] {
        border: 1px solid #d9e2e0;
        border-radius: 8px;
        background: #ffffff;
        padding: 14px;
      }

      div[data-testid="stFileUploader"] section {
        border-color: #d5dfdc;
        border-radius: 8px;
        background: #fbfcfd;
      }

      label,
      div[data-testid="stWidgetLabel"] p,
      div[data-testid="stWidgetLabel"] {
        font-size: 16px !important;
        color: #1e293b !important;
        font-weight: 700 !important;
        line-height: 1.4 !important;
      }

      .stMarkdown,
      .stAlert,
      .stCaptionContainer {
        font-size: 14px;
      }

      .stButton > button {
        border-radius: 8px;
        min-height: 44px;
        font-size: 16px;
        font-weight: 700;
        background-color: #334155 !important;
        color: white !important;
        border-color: #334155 !important;
      }
      .stButton > button:hover:not(:disabled) {
        background-color: #1e293b !important;
        border-color: #1e293b !important;
        color: white !important;
      }
      .stButton > button:disabled,
      .stDownloadButton > button:disabled {
        background-color: #e2e8f0 !important;
        color: #94a3b8 !important;
        border-color: #cbd5e1 !important;
      }

      .stDownloadButton > button {
        border-radius: 8px;
        min-height: 44px;
        font-size: 16px;
        font-weight: 700;
        background-color: #0d9488 !important;
        color: white !important;
        border-color: #0d9488 !important;
        width: 100%;
      }
      .stDownloadButton > button:hover:not(:disabled) {
        background-color: #0f766e !important;
        border-color: #0f766e !important;
        color: white !important;
      }

      div[data-baseweb="select"] {
        font-size: 14px;
      }

      @media (max-width: 560px) {
        .hero h1 {
          font-size: 26px;
        }
      }
    </style>
    """,
    unsafe_allow_html=True,
)


def uploaded_bytes(uploaded_file) -> bytes:
    data = uploaded_file.getvalue()
    if not data:
        raise ValueError("업로드한 파일이 비어 있습니다.")
    return data


@st.cache_data(show_spinner=False)
def sanitize_phone_numbers(low_data: bytes) -> tuple[bytes, int]:
    workbook = load_workbook(BytesIO(low_data))
    removed_count = 0

    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value

                if isinstance(value, (int, float)) and float(value).is_integer():
                    digits = str(int(value))
                    candidates = [digits, f"0{digits}"]
                    if any(PHONE_PATTERN.fullmatch(candidate) for candidate in candidates):
                        cell.value = None
                        removed_count += 1
                    continue

                if not isinstance(value, str):
                    continue

                matches = PHONE_PATTERN.findall(value)
                if not matches:
                    continue

                cleaned_value = PHONE_PATTERN.sub("", value)
                cleaned_value = re.sub(r"\s{2,}", " ", cleaned_value).strip(" -_/,.")
                cell.value = cleaned_value or None
                removed_count += len(matches)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), removed_count


@st.cache_data(show_spinner=False)
def available_dates_from_low(low_data: bytes) -> tuple[list[date], date | None]:
    workbook = load_workbook(BytesIO(low_data), data_only=True)
    sheet = workbook.active
    date_columns = find_date_columns(sheet)
    available_dates = sorted(date_columns)
    default_date = latest_populated_date(sheet, date_columns) if available_dates else None
    return available_dates, default_date



def generate_files(low_data: bytes, selected_dates: list[date]) -> list[tuple[str, bytes]]:
    missing_templates = [template_path.name for _, template_path in DEFAULT_TEMPLATE_FILES if not template_path.exists()]
    if missing_templates:
        missing_text = ", ".join(missing_templates)
        raise FileNotFoundError(f"앱 폴더에 기본 양식 파일이 없습니다: {missing_text}")

    with tempfile.TemporaryDirectory() as temp_dir_text:
        temp_dir = Path(temp_dir_text)
        low_path = temp_dir / "low.xlsx"
        low_path.write_bytes(low_data)

        output_files: list[tuple[str, bytes]] = []
        target_dates = selected_dates or None

        for template_label, template_path in DEFAULT_TEMPLATE_FILES:
            temp_output_path = temp_dir / f"{template_label}_result.xlsx"
            summary = fill_worklog_dates(
                low_file=low_path,
                template_file=template_path,
                output_file=temp_output_path,
                target_dates=target_dates,
            )
            date_part = "_".join(summary["target_dates"])
            output_name = f"{template_label}_result_{date_part}.xlsx"
            output_files.append((output_name, temp_output_path.read_bytes()))

    return output_files


def make_result_zip(output_files: list[tuple[str, bytes]]) -> tuple[str, bytes]:
    if not output_files:
        raise ValueError("압축할 결과 파일이 없습니다.")

    date_part = ""
    first_name = output_files[0][0]
    if "_result_" in first_name:
        date_part = first_name.rsplit("_result_", 1)[1].removesuffix(".xlsx")

    zip_name = f"worklog_results_{date_part or 'download'}.zip"
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        for output_name, output_data in output_files:
            zip_file.writestr(output_name, output_data)

    return zip_name, zip_buffer.getvalue()


if "result_zip" not in st.session_state:
    st.session_state.result_zip = None


st.markdown(
    """
    <section class="hero">
      <p class="eyebrow">WORKLOG AUTOMATION</p>
      <h1>업무일지 자동 생성</h1>
      <p>출결 파일을 올리고 작업일을 선택하면 결과 엑셀 2개를 ZIP으로 내려받을 수 있습니다.</p>
    </section>
    """,
    unsafe_allow_html=True,
)

with st.container():
    st.markdown("<p class='field-label'>출결 파일</p>", unsafe_allow_html=True)
    st.markdown("<p class='section-note'>업로드된 원본 파일은 서버에 저장되지 않으며, 포함된 전화번호는 개인정보 보호를 위해 자동 삭제됩니다.</p>", unsafe_allow_html=True)
    low_file = st.file_uploader("출결 파일 업로드", type=["xlsx"], label_visibility="collapsed")

    low_data: bytes | None = None
    selected_dates: list[date] = []

    if low_file is not None:
        try:
            raw_low_data = uploaded_bytes(low_file)
            sanitized_low_data, phone_count = sanitize_phone_numbers(raw_low_data)
            low_data = sanitized_low_data
            available_dates, default_date = available_dates_from_low(low_data)
            
            if phone_count > 0:
                st.info(f"🔒 개인정보 보호를 위해 엑셀 내 전화번호(추정) {phone_count}개를 자동 삭제했습니다.")
                
        except Exception as exc:
            st.error(f"출결 파일을 읽을 수 없습니다: {exc}")
            available_dates = []
            default_date = None

        st.markdown("<hr style='margin: 32px 0 24px; border: none; border-top: 1px solid #e2e8f0;' />", unsafe_allow_html=True)

        if available_dates:
            default_selection = [default_date] if default_date else []
            st.markdown("<p class='field-label'>생성 대상일</p>", unsafe_allow_html=True)
            st.markdown("<p class='section-note'>목록에서 기준일로 삼을 날짜를 여러 개 고를 수 있습니다. (기본 양식: worklog_set1, worklog_set2 사용)</p>", unsafe_allow_html=True)
            selected_dates = st.multiselect(
                "생성 대상일",
                options=available_dates,
                default=default_selection,
                format_func=lambda day: day.isoformat(),
                label_visibility="collapsed",
            )
        elif low_data is not None:
            st.warning("업로드된 파일에서 유효한 날짜 상태를 찾지 못했습니다. 엑셀 양식을 확인해 주세요.")
    else:
        st.info("먼저 출결 파일을 올려주세요.")

    st.write("")

    can_generate = low_data is not None
    button_label = "생성" if can_generate else "출결 파일을 먼저 올려주세요"
    col1, col2 = st.columns(2)

    with col1:
        if st.button(button_label, type="primary", use_container_width=True, disabled=not can_generate):
            try:
                output_files = generate_files(
                    low_data=low_data or b"",
                    selected_dates=sorted(selected_dates),
                )
            except Exception as exc:
                st.error(f"처리 중 오류가 발생했습니다: {exc}")
            else:
                zip_name, zip_data = make_result_zip(output_files)
                st.session_state.result_zip = {
                    "name": zip_name,
                    "data": zip_data,
                    "files": [output_name for output_name, _ in output_files],
                }

    with col2:
        is_ready = st.session_state.result_zip is not None
        download_data = st.session_state.result_zip["data"] if is_ready else b""
        download_name = st.session_state.result_zip["name"] if is_ready else "download.zip"
        
        st.download_button(
            label="다운로드",
            data=download_data,
            file_name=download_name,
            mime="application/zip",
            type="primary",
            use_container_width=True,
            disabled=not is_ready,
        )

    if st.session_state.result_zip:
        result_zip = st.session_state.result_zip
        files_html = "".join(f"<li>{output_name}</li>" for output_name in result_zip["files"])
        st.markdown(
            f"""
            <section class="result-panel">
              <h3>생성 완료</h3>
              <p>아래 파일들이 ZIP에 포함됩니다.</p>
              <ul class="file-list">{files_html}</ul>
            </section>
            """,
            unsafe_allow_html=True,
        )
